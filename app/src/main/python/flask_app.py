# -*- coding: utf-8 -*-
"""
==================================================================
نظام إدارة مدرسة الهدى الثانوية بنين - School Management System (ملف واحد فقط)
==================================================================
تطبيق Flask كامل بملف بايثون واحد (نماذج قاعدة البيانات + المسارات +
كل قوالب HTML مدمجة كنصوص) - يكفي تثبيت المتطلبات وتشغيل هذا الملف.

التشغيل:
    pip install flask flask_sqlalchemy werkzeug reportlab arabic_reshaper python-bidi openpyxl
    python flask_app.py
    ثم افتح: http://127.0.0.1:5000

لطباعة تقرير نتيجة الطالب (PDF) بحروف عربية صحيحة، ضع ملف الخط
"Amiri-Regular.ttf" بجانب هذا الملف (يمكن تحميله مجاناً من:
https://fonts.google.com/specimen/Amiri)

بيانات الدخول الافتراضية:
    اسم المستخدم: admin
    كلمة المرور : admin123
    (يُفضّل تغييرها بعد أول دخول من قسم "المستخدمون")
"""
import os
import io
import sys
from datetime import date, datetime
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, abort, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from jinja2 import DictLoader

# مكتبات تصدير PDF (نتيجة الطالب) - reportlab + دعم النص العربي
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

# مكتبة إكسل (استيراد / تصدير بيانات الطلاب)
import openpyxl
from openpyxl.utils import get_column_letter

SCHOOL_NAME = "مدرسة الهدى الثانوية بنين"

# عند تحويل البرنامج إلى ملف تنفيذي واحد (.exe) عبر PyInstaller، تكون قيمة
# __file__ داخل مجلد مؤقت يُحذف بعد إغلاق البرنامج. لذلك نحفظ قاعدة البيانات
# بجانب ملف .exe نفسه حتى تبقى بياناتك محفوظة بين مرة وأخرى.
if os.environ.get('SCHOOL_DATA_DIR'):
    BASE_DIR = os.environ['SCHOOL_DATA_DIR']
elif getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

os.makedirs(BASE_DIR, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'school.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ---------------------------------------------------------------------------
# إعداد الخط العربي لملفات PDF
# ضع ملف الخط "Amiri-Regular.ttf" بجانب هذا الملف حتى تظهر النصوص العربية
# بشكل صحيح داخل تقرير النتيجة (PDF). يمكن تحميله مجاناً من Google Fonts:
# https://fonts.google.com/specimen/Amiri
# في حال عدم توفر الخط، سيتم استخدام خط افتراضي وقد لا تظهر الحروف العربية
# بشكل صحيح داخل ملف الـ PDF (الصفحات داخل الموقع نفسه لا تتأثر بهذا إطلاقاً).
# ---------------------------------------------------------------------------
ARABIC_FONT_NAME = 'Amiri'
ARABIC_FONT_PATH = os.path.join(BASE_DIR, 'Amiri-Regular.ttf')
ARABIC_FONT_READY = False
try:
    if os.path.exists(ARABIC_FONT_PATH):
        pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, ARABIC_FONT_PATH))
        ARABIC_FONT_READY = True
except Exception:
    ARABIC_FONT_READY = False

PDF_FONT = ARABIC_FONT_NAME if ARABIC_FONT_READY else 'Helvetica'


def ar(text):
    """يهيئ النص العربي (يصل الحروف ويرتبها من اليمين لليسار) ليظهر بشكل
    صحيح داخل تقارير PDF. يُستخدم لأي نص عربي يُكتب داخل ملف PDF فقط."""
    if text is None:
        return ''
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)

# ---------------------------------------------------------------------------
# القوالب (Templates) - مدمجة داخل الملف كنصوص باستخدام DictLoader
# ---------------------------------------------------------------------------
TEMPLATES = {
    "attendance.html": """{% extends 'base.html' %}
{% block title %}الحضور والغياب{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-calendar-check-fill"></i> الحضور والغياب</h3>

<form method="get" class="card p-3 mb-3">
  <div class="row g-2">
    <div class="col-md-5">
      <label class="form-label">الصف</label>
      <select name="class_id" class="form-select" onchange="this.form.submit()">
        <option value="">اختر الصف</option>
        {% for c in classes %}
        <option value="{{ c.id }}" {% if selected_class|string == c.id|string %}selected{% endif %}>{{ c.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-4">
      <label class="form-label">التاريخ</label>
      <input type="date" name="date" value="{{ selected_date }}" class="form-control" onchange="this.form.submit()">
    </div>
  </div>
</form>

{% if students %}
<form method="post">
  <input type="hidden" name="class_id" value="{{ selected_class }}">
  <input type="hidden" name="date" value="{{ selected_date }}">
  <div class="card">
    <table class="table mb-0">
      <thead><tr><th>الطالب</th><th>الحالة</th></tr></thead>
      <tbody>
        {% for s in students %}
        <tr>
          <td>{{ s.full_name }}</td>
          <td>
            <select name="status_{{ s.id }}" class="form-select form-select-sm w-auto d-inline-block">
              {% set current = existing_map.get(s.id, 'حاضر') %}
              <option value="حاضر" {% if current=='حاضر' %}selected{% endif %}>حاضر</option>
              <option value="غائب" {% if current=='غائب' %}selected{% endif %}>غائب</option>
              <option value="متأخر" {% if current=='متأخر' %}selected{% endif %}>متأخر</option>
              <option value="غياب بعذر" {% if current=='غياب بعذر' %}selected{% endif %}>غياب بعذر</option>
            </select>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  <button class="btn btn-primary mt-3"><i class="bi bi-check-lg"></i> حفظ الحضور</button>
</form>
{% elif selected_class %}
<p class="text-muted">لا يوجد طلاب في هذا الصف</p>
{% else %}
<p class="text-muted">الرجاء اختيار الصف لعرض قائمة الطلاب</p>
{% endif %}
{% endblock %}
""",
    "base.html": """<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{% block title %}""" + SCHOOL_NAME + """{% endblock %}</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.rtl.min.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  :root { --base-font-size: 16px; }
  html { font-size: var(--base-font-size); }
  body { font-family: 'Cairo', sans-serif; background:#f4f6f9; font-size: 1rem; }
  .navbar-brand { font-weight:700; }
  .sidebar { min-height: calc(100vh - 56px); background:#1e2a3a; }
  .sidebar a { color:#c9d3e0; display:block; padding:.65rem 1.2rem; text-decoration:none; border-right:3px solid transparent; }
  .sidebar a:hover, .sidebar a.active { background:#25344a; color:#fff; border-right-color:#4e8cff; }
  .card-stat { border:none; border-radius:14px; color:#fff; }
  .card-stat .stat-num { font-size:2rem; font-weight:700; }
  .table thead { background:#eef1f6; }
  .brand-badge { background:#4e8cff; color:#fff; border-radius:8px; padding:.15rem .5rem; font-size:.8rem; }
  .font-size-controls { display:flex; align-items:center; gap:.35rem; }
  .font-size-controls button { width:30px; height:30px; padding:0; line-height:1; border-radius:6px; }
</style>
</head>
<body>
<nav class="navbar navbar-dark bg-dark px-3">
  <a class="navbar-brand" href="{{ url_for('dashboard') }}"><i class="bi bi-mortarboard-fill"></i> """ + SCHOOL_NAME + """</a>
  <div class="d-flex align-items-center gap-3">
    <div class="font-size-controls">
      <button type="button" class="btn btn-sm btn-outline-light" onclick="changeFontSize(-1)" title="تصغير الخط">A-</button>
      <button type="button" class="btn btn-sm btn-outline-light" onclick="changeFontSize(1)" title="تكبير الخط">A+</button>
    </div>
    {% if session.get('user_id') %}
    <span class="text-light small">{{ session.get('full_name') }} <span class="brand-badge">{{ 'مدير' if session.get('role')=='admin' else 'معلم' }}</span></span>
    <a href="{{ url_for('logout') }}" class="btn btn-sm btn-outline-light"><i class="bi bi-box-arrow-right"></i> تسجيل الخروج</a>
    {% endif %}
  </div>
</nav>

<div class="d-flex">
  {% if session.get('user_id') %}
  <div class="sidebar" style="width:230px;">
    <a href="{{ url_for('dashboard') }}"><i class="bi bi-speedometer2 ms-1"></i> لوحة التحكم</a>
    <a href="{{ url_for('students_list') }}"><i class="bi bi-people-fill ms-1"></i> الطلاب</a>
    <a href="{{ url_for('teachers_list') }}"><i class="bi bi-person-badge-fill ms-1"></i> المعلمون</a>
    <a href="{{ url_for('classes_list') }}"><i class="bi bi-door-open-fill ms-1"></i> الصفوف</a>
    <a href="{{ url_for('subjects_list') }}"><i class="bi bi-book-fill ms-1"></i> المواد الدراسية</a>
    <a href="{{ url_for('attendance_page') }}"><i class="bi bi-calendar-check-fill ms-1"></i> الحضور والغياب</a>
    <a href="{{ url_for('grades_page') }}"><i class="bi bi-clipboard-data-fill ms-1"></i> الدرجات</a>
    <a href="{{ url_for('fees_list') }}"><i class="bi bi-cash-coin ms-1"></i> الرسوم الدراسية</a>
    {% if session.get('role') == 'admin' %}
    <a href="{{ url_for('users_list') }}"><i class="bi bi-shield-lock-fill ms-1"></i> المستخدمون</a>
    {% endif %}
  </div>
  {% endif %}

  <div class="flex-grow-1 p-4">
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
            {{ message }}
            <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
          </div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    {% block content %}{% endblock %}
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script>
  (function () {
    var MIN_SIZE = 12, MAX_SIZE = 24, STEP = 1, STORAGE_KEY = 'huda_font_size';
    var root = document.documentElement;

    function applySize(size) {
      root.style.setProperty('--base-font-size', size + 'px');
    }

    var saved = parseInt(localStorage.getItem(STORAGE_KEY), 10);
    if (!isNaN(saved)) applySize(saved);

    window.changeFontSize = function (direction) {
      var current = parseInt(getComputedStyle(root).getPropertyValue('--base-font-size'), 10) || 16;
      var next = Math.max(MIN_SIZE, Math.min(MAX_SIZE, current + direction * STEP));
      applySize(next);
      localStorage.setItem(STORAGE_KEY, next);
    };
  })();
</script>
</body>
</html>
""",
    "class_form.html": """{% extends 'base.html' %}
{% block title %}{{ 'تعديل صف' if school_class else 'إضافة صف' }}{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-door-open"></i> {{ 'تعديل صف' if school_class else 'إضافة صف جديد' }}</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">اسم الصف *</label>
      <input type="text" name="name" class="form-control" required placeholder="مثال: الصف الأول أ" value="{{ school_class.name if school_class else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">المرحلة *</label>
      <input type="text" name="grade_level" class="form-control" required placeholder="مثال: الصف الأول" value="{{ school_class.grade_level if school_class else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">العام الدراسي</label>
      <input type="text" name="academic_year" class="form-control" value="{{ school_class.academic_year if school_class else '2025-2026' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">معلم الصف</label>
      <select name="homeroom_teacher_id" class="form-select">
        <option value="">بدون</option>
        {% for t in teachers %}
        <option value="{{ t.id }}" {% if school_class and school_class.homeroom_teacher_id==t.id %}selected{% endif %}>{{ t.full_name }}</option>
        {% endfor %}
      </select>
    </div>
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('classes_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "classes.html": """{% extends 'base.html' %}
{% block title %}الصفوف{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-door-open-fill"></i> الصفوف الدراسية</h3>
  {% if session.get('role') == 'admin' %}
  <a href="{{ url_for('class_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> إضافة صف</a>
  {% endif %}
</div>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>اسم الصف</th><th>المرحلة</th><th>العام الدراسي</th><th>معلم الصف</th><th>عدد الطلاب</th><th></th></tr></thead>
    <tbody>
      {% for c in classes %}
      <tr>
        <td><a href="{{ url_for('report_class', class_id=c.id) }}">{{ c.name }}</a></td>
        <td>{{ c.grade_level }}</td>
        <td>{{ c.academic_year }}</td>
        <td>{{ c.homeroom_teacher.full_name if c.homeroom_teacher else '-' }}</td>
        <td>{{ c.student_count }}</td>
        <td class="text-nowrap">
          {% if session.get('role') == 'admin' %}
          <a href="{{ url_for('class_edit', class_id=c.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-pencil"></i></a>
          <form method="post" action="{{ url_for('class_delete', class_id=c.id) }}" class="d-inline" onsubmit="return confirm('تأكيد حذف الصف؟');">
            <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
          </form>
          {% endif %}
        </td>
      </tr>
      {% else %}
      <tr><td colspan="6" class="text-center text-muted py-4">لا توجد صفوف</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "dashboard.html": """{% extends 'base.html' %}
{% block title %}لوحة التحكم{% endblock %}
{% block content %}
<h3 class="mb-4"><i class="bi bi-speedometer2"></i> لوحة التحكم</h3>

<div class="row g-3 mb-4">
  <div class="col-md-3">
    <div class="card card-stat p-3" style="background:linear-gradient(135deg,#4e8cff,#2a5fd8);">
      <div class="d-flex justify-content-between align-items-center">
        <div><div class="stat-num">{{ stats.students }}</div><div>الطلاب</div></div>
        <i class="bi bi-people-fill" style="font-size:2.2rem; opacity:.6;"></i>
      </div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card card-stat p-3" style="background:linear-gradient(135deg,#22c07a,#128a54);">
      <div class="d-flex justify-content-between align-items-center">
        <div><div class="stat-num">{{ stats.teachers }}</div><div>المعلمون</div></div>
        <i class="bi bi-person-badge-fill" style="font-size:2.2rem; opacity:.6;"></i>
      </div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card card-stat p-3" style="background:linear-gradient(135deg,#f5a623,#c9790c);">
      <div class="d-flex justify-content-between align-items-center">
        <div><div class="stat-num">{{ stats.classes }}</div><div>الصفوف</div></div>
        <i class="bi bi-door-open-fill" style="font-size:2.2rem; opacity:.6;"></i>
      </div>
    </div>
  </div>
  <div class="col-md-3">
    <div class="card card-stat p-3" style="background:linear-gradient(135deg,#a55eea,#7b2ff7);">
      <div class="d-flex justify-content-between align-items-center">
        <div><div class="stat-num">{{ stats.subjects }}</div><div>المواد الدراسية</div></div>
        <i class="bi bi-book-fill" style="font-size:2.2rem; opacity:.6;"></i>
      </div>
    </div>
  </div>
</div>

<div class="row g-3">
  <div class="col-md-4">
    <div class="card p-3">
      <h6><i class="bi bi-calendar-check"></i> حضور اليوم ({{ today }})</h6>
      <p class="mb-1">إجمالي السجلات: <strong>{{ today_attendance }}</strong></p>
      <p class="mb-0 text-danger">الغياب: <strong>{{ today_absent }}</strong></p>
    </div>
  </div>
  <div class="col-md-8">
    <div class="card p-3">
      <h6><i class="bi bi-person-plus-fill"></i> آخر الطلاب المسجلين</h6>
      <table class="table table-sm mb-0">
        <thead><tr><th>الاسم</th><th>الصف</th><th></th></tr></thead>
        <tbody>
          {% for s in recent_students %}
          <tr>
            <td>{{ s.full_name }}</td>
            <td>{{ s.school_class.name if s.school_class else '-' }}</td>
            <td><a href="{{ url_for('student_profile', student_id=s.id) }}" class="btn btn-sm btn-outline-primary">عرض</a></td>
          </tr>
          {% else %}
          <tr><td colspan="3" class="text-center text-muted">لا يوجد طلاب بعد</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
</div>

<div class="card p-3 mt-3">
  <h6><i class="bi bi-door-open"></i> الصفوف الدراسية</h6>
  <div class="row g-2">
    {% for c in classes %}
    <div class="col-md-3">
      <a href="{{ url_for('report_class', class_id=c.id) }}" class="text-decoration-none">
        <div class="border rounded p-2 text-center">
          <div class="fw-bold">{{ c.name }}</div>
          <div class="small text-muted">{{ c.student_count }} طالب</div>
        </div>
      </a>
    </div>
    {% else %}
    <p class="text-muted">لا توجد صفوف بعد</p>
    {% endfor %}
  </div>
</div>
{% endblock %}
""",
    "fee_form.html": """{% extends 'base.html' %}
{% block title %}تسجيل دفعة{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-cash-coin"></i> تسجيل دفعة رسوم</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">الطالب *</label>
      <select name="student_id" class="form-select" required>
        <option value="">اختر الطالب</option>
        {% for s in students %}
        <option value="{{ s.id }}">{{ s.full_name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-6">
      <label class="form-label">المبلغ *</label>
      <input type="number" step="0.01" name="amount" class="form-control" required>
    </div>
    <div class="col-md-6">
      <label class="form-label">تاريخ الدفع</label>
      <input type="date" name="payment_date" class="form-control">
    </div>
    <div class="col-md-6">
      <label class="form-label">ملاحظات</label>
      <input type="text" name="description" class="form-control" placeholder="مثال: رسوم الفصل الأول">
    </div>
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('fees_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "fees.html": """{% extends 'base.html' %}
{% block title %}الرسوم الدراسية{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-cash-coin"></i> الرسوم الدراسية</h3>
  {% if session.get('role') == 'admin' %}
  <a href="{{ url_for('fee_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> تسجيل دفعة</a>
  {% endif %}
</div>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>الطالب</th><th>المبلغ</th><th>تاريخ الدفع</th><th>ملاحظات</th></tr></thead>
    <tbody>
      {% for p in payments %}
      <tr>
        <td><a href="{{ url_for('student_profile', student_id=p.student_id) }}">{{ p.student.full_name }}</a></td>
        <td>{{ p.amount }}</td>
        <td>{{ p.payment_date }}</td>
        <td>{{ p.description or '-' }}</td>
      </tr>
      {% else %}
      <tr><td colspan="4" class="text-center text-muted py-4">لا توجد دفعات مسجلة</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "grades.html": """{% extends 'base.html' %}
{% block title %}الدرجات{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-clipboard-data-fill"></i> الدرجات</h3>

<form method="get" class="card p-3 mb-3">
  <div class="row g-2">
    <div class="col-md-3">
      <label class="form-label">الصف</label>
      <select name="class_id" class="form-select">
        <option value="">اختر الصف</option>
        {% for c in classes %}
        <option value="{{ c.id }}" {% if selected_class|string == c.id|string %}selected{% endif %}>{{ c.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">المادة</label>
      <select name="subject_id" class="form-select">
        <option value="">اختر المادة</option>
        {% for s in subjects %}
        <option value="{{ s.id }}" {% if selected_subject|string == s.id|string %}selected{% endif %}>{{ s.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">الفصل الدراسي</label>
      <select name="term" class="form-select">
        <option value="الفصل الأول" {% if term=='الفصل الأول' %}selected{% endif %}>الفصل الأول</option>
        <option value="الفصل الثاني" {% if term=='الفصل الثاني' %}selected{% endif %}>الفصل الثاني</option>
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">نوع التقييم</label>
      <select name="exam_type" class="form-select">
        <option value="اختبار قصير" {% if exam_type=='اختبار قصير' %}selected{% endif %}>اختبار قصير</option>
        <option value="منتصف الفصل" {% if exam_type=='منتصف الفصل' %}selected{% endif %}>منتصف الفصل</option>
        <option value="اختبار نهائي" {% if exam_type=='اختبار نهائي' %}selected{% endif %}>اختبار نهائي</option>
      </select>
    </div>
  </div>
  <button class="btn btn-outline-secondary mt-3"><i class="bi bi-search"></i> عرض</button>
</form>

{% if students %}
<form method="post">
  <input type="hidden" name="class_id" value="{{ selected_class }}">
  <input type="hidden" name="subject_id" value="{{ selected_subject }}">
  <input type="hidden" name="term" value="{{ term }}">
  <input type="hidden" name="exam_type" value="{{ exam_type }}">
  <div class="row mb-2">
    <div class="col-md-3">
      <label class="form-label">الدرجة القصوى</label>
      <input type="number" step="0.01" name="max_score" class="form-control" value="{{ existing_map.values()|map(attribute='max_score')|first or 100 }}">
    </div>
  </div>
  <div class="card">
    <table class="table mb-0">
      <thead><tr><th>الطالب</th><th>الدرجة</th></tr></thead>
      <tbody>
        {% for s in students %}
        <tr>
          <td>{{ s.full_name }}</td>
          <td>
            <input type="number" step="0.01" name="score_{{ s.id }}" class="form-control form-control-sm" style="width:120px;"
                   value="{{ existing_map[s.id].score if s.id in existing_map else '' }}">
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  <button class="btn btn-primary mt-3"><i class="bi bi-check-lg"></i> حفظ الدرجات</button>
</form>
{% elif selected_class and selected_subject %}
<p class="text-muted">لا يوجد طلاب في هذا الصف</p>
{% else %}
<p class="text-muted">الرجاء اختيار الصف والمادة لعرض قائمة الطلاب</p>
{% endif %}
{% endblock %}
""",
    "login.html": """<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>تسجيل الدخول - """ + SCHOOL_NAME + """</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.rtl.min.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  body { font-family:'Cairo',sans-serif; background:linear-gradient(135deg,#1e2a3a,#4e8cff); min-height:100vh; display:flex; align-items:center; }
  .login-card { border:none; border-radius:18px; box-shadow:0 20px 50px rgba(0,0,0,.25); overflow:hidden; }
  .login-header { background:#1e2a3a; color:#fff; padding:2rem; text-align:center; }
</style>
</head>
<body>
<div class="container">
  <div class="row justify-content-center">
    <div class="col-md-5 col-lg-4">
      <div class="card login-card">
        <div class="login-header">
          <i class="bi bi-mortarboard-fill" style="font-size:2.5rem;"></i>
          <h4 class="mt-2 mb-0">""" + SCHOOL_NAME + """</h4>
        </div>
        <div class="card-body p-4">
          {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
              {% for category, message in messages %}
                <div class="alert alert-{{ category }}">{{ message }}</div>
              {% endfor %}
            {% endif %}
          {% endwith %}
          <form method="post">
            <div class="mb-3">
              <label class="form-label">اسم المستخدم</label>
              <input type="text" name="username" class="form-control" required autofocus>
            </div>
            <div class="mb-3">
              <label class="form-label">كلمة المرور</label>
              <input type="password" name="password" class="form-control" required>
            </div>
            <button type="submit" class="btn btn-primary w-100 py-2">
              <i class="bi bi-box-arrow-in-left"></i> تسجيل الدخول
            </button>
          </form>
          <p class="text-muted text-center mt-3 small">الحساب الافتراضي: admin / admin123</p>
        </div>
      </div>
    </div>
  </div>
</div>
</body>
</html>
""",
    "report_class.html": """{% extends 'base.html' %}
{% block title %}تقرير الصف{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-file-earmark-bar-graph"></i> تقرير صف: {{ school_class.name }}</h3>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>الطالب</th><th>متوسط الدرجات %</th><th>أيام الحضور</th><th>أيام الغياب</th></tr></thead>
    <tbody>
      {% for row in rows %}
      <tr>
        <td><a href="{{ url_for('student_profile', student_id=row.student.id) }}">{{ row.student.full_name }}</a></td>
        <td>
          {% if row.avg is not none %}
            <span class="badge bg-{{ 'success' if row.avg >= 60 else 'danger' }}">{{ row.avg }}%</span>
          {% else %}
            <span class="text-muted">لا يوجد</span>
          {% endif %}
        </td>
        <td class="text-success">{{ row.present }}</td>
        <td class="text-danger">{{ row.absent }}</td>
      </tr>
      {% else %}
      <tr><td colspan="4" class="text-center text-muted py-4">لا يوجد طلاب في هذا الصف</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "student_form.html": """{% extends 'base.html' %}
{% block title %}{{ 'تعديل طالب' if student else 'إضافة طالب' }}{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-person-plus"></i> {{ 'تعديل بيانات طالب' if student else 'إضافة طالب جديد' }}</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">الاسم الكامل *</label>
      <input type="text" name="full_name" class="form-control" required value="{{ student.full_name if student else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">الرقم الوطني / رقم الهوية</label>
      <input type="text" name="national_id" class="form-control" value="{{ student.national_id if student else '' }}">
    </div>
    <div class="col-md-4">
      <label class="form-label">الجنس</label>
      <select name="gender" class="form-select">
        <option value="ذكر" {% if student and student.gender=='ذكر' %}selected{% endif %}>ذكر</option>
        <option value="أنثى" {% if student and student.gender=='أنثى' %}selected{% endif %}>أنثى</option>
      </select>
    </div>
    <div class="col-md-4">
      <label class="form-label">تاريخ الميلاد</label>
      <input type="date" name="birth_date" class="form-control" value="{{ student.birth_date if student else '' }}">
    </div>
    <div class="col-md-4">
      <label class="form-label">الصف</label>
      <select name="class_id" class="form-select">
        <option value="">بدون صف</option>
        {% for c in classes %}
        <option value="{{ c.id }}" {% if student and student.class_id==c.id %}selected{% endif %}>{{ c.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-6">
      <label class="form-label">العنوان</label>
      <input type="text" name="address" class="form-control" value="{{ student.address if student else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">هاتف الطالب</label>
      <input type="text" name="phone" class="form-control" value="{{ student.phone if student else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">اسم ولي الأمر</label>
      <input type="text" name="parent_name" class="form-control" value="{{ student.parent_name if student else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">هاتف ولي الأمر</label>
      <input type="text" name="parent_phone" class="form-control" value="{{ student.parent_phone if student else '' }}">
    </div>
    {% if not student %}
    <div class="col-md-6">
      <label class="form-label">تاريخ التسجيل</label>
      <input type="date" name="enrollment_date" class="form-control">
    </div>
    {% endif %}
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('students_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "student_profile.html": """{% extends 'base.html' %}
{% block title %}ملف الطالب{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-person-vcard"></i> {{ s.full_name }}</h3>
  <div>
    <a href="{{ url_for('student_result', student_id=s.id) }}" class="btn btn-outline-success"><i class="bi bi-file-earmark-bar-graph"></i> نتيجة الطالب</a>
    <a href="{{ url_for('student_edit', student_id=s.id) }}" class="btn btn-outline-primary"><i class="bi bi-pencil"></i> تعديل</a>
  </div>
</div>

<div class="row g-3 mb-3">
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold">{{ s.school_class.name if s.school_class else '-' }}</div><div class="text-muted small">الصف</div></div></div>
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold text-success">{{ total_present }}</div><div class="text-muted small">أيام حضور</div></div></div>
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold text-danger">{{ total_absent }}</div><div class="text-muted small">أيام غياب</div></div></div>
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold">{{ total_paid }}</div><div class="text-muted small">إجمالي المدفوع</div></div></div>
</div>

<div class="row g-3">
  <div class="col-md-6">
    <div class="card p-3">
      <h6><i class="bi bi-person-lines-fill"></i> البيانات الشخصية</h6>
      <table class="table table-sm">
        <tr><th>الرقم الوطني</th><td>{{ s.national_id or '-' }}</td></tr>
        <tr><th>الجنس</th><td>{{ s.gender or '-' }}</td></tr>
        <tr><th>تاريخ الميلاد</th><td>{{ s.birth_date or '-' }}</td></tr>
        <tr><th>العنوان</th><td>{{ s.address or '-' }}</td></tr>
        <tr><th>هاتف الطالب</th><td>{{ s.phone or '-' }}</td></tr>
        <tr><th>ولي الأمر</th><td>{{ s.parent_name or '-' }}</td></tr>
        <tr><th>هاتف ولي الأمر</th><td>{{ s.parent_phone or '-' }}</td></tr>
      </table>
    </div>
  </div>
  <div class="col-md-6">
    <div class="card p-3 mb-3">
      <h6><i class="bi bi-clipboard-data"></i> الدرجات</h6>
      <table class="table table-sm">
        <thead><tr><th>المادة</th><th>الفصل</th><th>النوع</th><th>الدرجة</th></tr></thead>
        <tbody>
          {% for g in grades %}
          <tr><td>{{ g.subject.name }}</td><td>{{ g.term }}</td><td>{{ g.exam_type }}</td><td>{{ g.score }}/{{ g.max_score }} ({{ g.percentage }}%)</td></tr>
          {% else %}
          <tr><td colspan="4" class="text-muted text-center">لا توجد درجات</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    <div class="card p-3">
      <h6><i class="bi bi-calendar-week"></i> آخر سجلات الحضور</h6>
      <table class="table table-sm">
        <thead><tr><th>التاريخ</th><th>الحالة</th></tr></thead>
        <tbody>
          {% for a in attendances %}
          <tr><td>{{ a.date }}</td><td>{{ a.status }}</td></tr>
          {% else %}
          <tr><td colspan="2" class="text-muted text-center">لا توجد سجلات</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
</div>
{% endblock %}
""",
    "students.html": """{% extends 'base.html' %}
{% block title %}الطلاب{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-people-fill"></i> الطلاب</h3>
  <div>
    <a href="{{ url_for('students_export') }}" class="btn btn-outline-success"><i class="bi bi-file-earmark-excel"></i> تصدير Excel</a>
    {% if session.get('role') == 'admin' %}
    <a href="{{ url_for('students_import') }}" class="btn btn-outline-secondary"><i class="bi bi-file-earmark-arrow-up"></i> استيراد من Excel</a>
    {% endif %}
    <a href="{{ url_for('student_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> إضافة طالب</a>
  </div>
</div>

<form method="get" class="card p-3 mb-3">
  <div class="row g-2">
    <div class="col-md-5">
      <input type="text" name="q" value="{{ q }}" class="form-control" placeholder="بحث بالاسم...">
    </div>
    <div class="col-md-4">
      <select name="class_id" class="form-select">
        <option value="">كل الصفوف</option>
        {% for c in classes %}
        <option value="{{ c.id }}" {% if selected_class|string == c.id|string %}selected{% endif %}>{{ c.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <button class="btn btn-outline-secondary w-100"><i class="bi bi-search"></i> بحث</button>
    </div>
  </div>
</form>

<div class="card">
  <table class="table table-hover mb-0">
    <thead>
      <tr>
        <th>الاسم</th><th>الجنس</th><th>الصف</th><th>هاتف ولي الأمر</th><th>الحالة</th><th></th>
      </tr>
    </thead>
    <tbody>
      {% for s in students %}
      <tr>
        <td><a href="{{ url_for('student_profile', student_id=s.id) }}">{{ s.full_name }}</a></td>
        <td>{{ s.gender or '-' }}</td>
        <td>{{ s.school_class.name if s.school_class else '-' }}</td>
        <td>{{ s.parent_phone or '-' }}</td>
        <td>
          {% if s.is_active %}<span class="badge bg-success">نشط</span>{% else %}<span class="badge bg-secondary">غير نشط</span>{% endif %}
        </td>
        <td class="text-nowrap">
          <a href="{{ url_for('student_edit', student_id=s.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-pencil"></i></a>
          <form method="post" action="{{ url_for('student_delete', student_id=s.id) }}" class="d-inline" onsubmit="return confirm('تأكيد حذف الطالب؟');">
            <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
          </form>
        </td>
      </tr>
      {% else %}
      <tr><td colspan="6" class="text-center text-muted py-4">لا يوجد طلاب</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "subject_form.html": """{% extends 'base.html' %}
{% block title %}{{ 'تعديل مادة' if subject else 'إضافة مادة' }}{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-book"></i> {{ 'تعديل مادة' if subject else 'إضافة مادة جديدة' }}</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">اسم المادة *</label>
      <input type="text" name="name" class="form-control" required value="{{ subject.name if subject else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">المرحلة</label>
      <input type="text" name="grade_level" class="form-control" value="{{ subject.grade_level if subject else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">المعلم المسؤول</label>
      <select name="teacher_id" class="form-select">
        <option value="">بدون</option>
        {% for t in teachers %}
        <option value="{{ t.id }}" {% if subject and subject.teacher_id==t.id %}selected{% endif %}>{{ t.full_name }}</option>
        {% endfor %}
      </select>
    </div>
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('subjects_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "subjects.html": """{% extends 'base.html' %}
{% block title %}المواد الدراسية{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-book-fill"></i> المواد الدراسية</h3>
  {% if session.get('role') == 'admin' %}
  <a href="{{ url_for('subject_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> إضافة مادة</a>
  {% endif %}
</div>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>اسم المادة</th><th>المرحلة</th><th>المعلم المسؤول</th><th></th></tr></thead>
    <tbody>
      {% for s in subjects %}
      <tr>
        <td>{{ s.name }}</td>
        <td>{{ s.grade_level or '-' }}</td>
        <td>{{ s.teacher.full_name if s.teacher else '-' }}</td>
        <td class="text-nowrap">
          {% if session.get('role') == 'admin' %}
          <a href="{{ url_for('subject_edit', subject_id=s.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-pencil"></i></a>
          <form method="post" action="{{ url_for('subject_delete', subject_id=s.id) }}" class="d-inline" onsubmit="return confirm('تأكيد حذف المادة؟');">
            <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
          </form>
          {% endif %}
        </td>
      </tr>
      {% else %}
      <tr><td colspan="4" class="text-center text-muted py-4">لا توجد مواد</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "teacher_form.html": """{% extends 'base.html' %}
{% block title %}{{ 'تعديل معلم' if teacher else 'إضافة معلم' }}{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-person-badge"></i> {{ 'تعديل بيانات معلم' if teacher else 'إضافة معلم جديد' }}</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">الاسم الكامل *</label>
      <input type="text" name="full_name" class="form-control" required value="{{ teacher.full_name if teacher else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">التخصص</label>
      <input type="text" name="specialty" class="form-control" value="{{ teacher.specialty if teacher else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">الهاتف</label>
      <input type="text" name="phone" class="form-control" value="{{ teacher.phone if teacher else '' }}">
    </div>
    <div class="col-md-6">
      <label class="form-label">البريد الإلكتروني</label>
      <input type="email" name="email" class="form-control" value="{{ teacher.email if teacher else '' }}">
    </div>
    {% if not teacher %}
    <div class="col-md-6">
      <label class="form-label">تاريخ التعيين</label>
      <input type="date" name="hire_date" class="form-control">
    </div>
    {% endif %}
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('teachers_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "teachers.html": """{% extends 'base.html' %}
{% block title %}المعلمون{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-person-badge-fill"></i> المعلمون</h3>
  {% if session.get('role') == 'admin' %}
  <a href="{{ url_for('teacher_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> إضافة معلم</a>
  {% endif %}
</div>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>الاسم</th><th>التخصص</th><th>الهاتف</th><th>البريد الإلكتروني</th><th>تاريخ التعيين</th><th></th></tr></thead>
    <tbody>
      {% for t in teachers %}
      <tr>
        <td>{{ t.full_name }}</td>
        <td>{{ t.specialty or '-' }}</td>
        <td>{{ t.phone or '-' }}</td>
        <td>{{ t.email or '-' }}</td>
        <td>{{ t.hire_date }}</td>
        <td class="text-nowrap">
          {% if session.get('role') == 'admin' %}
          <a href="{{ url_for('teacher_edit', teacher_id=t.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-pencil"></i></a>
          <form method="post" action="{{ url_for('teacher_delete', teacher_id=t.id) }}" class="d-inline" onsubmit="return confirm('تأكيد حذف المعلم؟');">
            <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
          </form>
          {% endif %}
        </td>
      </tr>
      {% else %}
      <tr><td colspan="6" class="text-center text-muted py-4">لا يوجد معلمون</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "user_form.html": """{% extends 'base.html' %}
{% block title %}إضافة مستخدم{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-person-plus"></i> إضافة مستخدم جديد</h3>
<div class="card p-4">
<form method="post">
  <div class="row g-3">
    <div class="col-md-6">
      <label class="form-label">اسم المستخدم *</label>
      <input type="text" name="username" class="form-control" required>
    </div>
    <div class="col-md-6">
      <label class="form-label">كلمة المرور *</label>
      <input type="password" name="password" class="form-control" required>
    </div>
    <div class="col-md-6">
      <label class="form-label">الاسم الكامل *</label>
      <input type="text" name="full_name" class="form-control" required>
    </div>
    <div class="col-md-6">
      <label class="form-label">الصلاحية</label>
      <select name="role" class="form-select">
        <option value="teacher">معلم</option>
        <option value="admin">مدير</option>
      </select>
    </div>
    <div class="col-md-6">
      <label class="form-label">ربط بمعلم (اختياري)</label>
      <select name="teacher_id" class="form-select">
        <option value="">بدون</option>
        {% for t in teachers %}
        <option value="{{ t.id }}">{{ t.full_name }}</option>
        {% endfor %}
      </select>
    </div>
  </div>
  <div class="mt-4">
    <button class="btn btn-primary"><i class="bi bi-check-lg"></i> حفظ</button>
    <a href="{{ url_for('users_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </div>
</form>
</div>
{% endblock %}
""",
    "users.html": """{% extends 'base.html' %}
{% block title %}المستخدمون{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-shield-lock-fill"></i> المستخدمون</h3>
  <a href="{{ url_for('user_add') }}" class="btn btn-primary"><i class="bi bi-plus-lg"></i> إضافة مستخدم</a>
</div>
<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>اسم المستخدم</th><th>الاسم الكامل</th><th>الصلاحية</th><th></th></tr></thead>
    <tbody>
      {% for u in users %}
      <tr>
        <td>{{ u.username }}</td>
        <td>{{ u.full_name }}</td>
        <td><span class="badge bg-{{ 'primary' if u.role=='admin' else 'secondary' }}">{{ 'مدير' if u.role=='admin' else 'معلم' }}</span></td>
        <td>
          <form method="post" action="{{ url_for('user_delete', user_id=u.id) }}" class="d-inline" onsubmit="return confirm('تأكيد حذف المستخدم؟');">
            <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
          </form>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "student_result.html": """{% extends 'base.html' %}
{% block title %}نتيجة الطالب{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3><i class="bi bi-file-earmark-bar-graph"></i> نتيجة الطالب: {{ s.full_name }}</h3>
  <div>
    <a href="{{ url_for('student_result_pdf', student_id=s.id, term=term) }}" class="btn btn-danger">
      <i class="bi bi-file-earmark-pdf"></i> طباعة PDF
    </a>
    <a href="{{ url_for('student_profile', student_id=s.id) }}" class="btn btn-outline-secondary">رجوع</a>
  </div>
</div>

<form method="get" class="card p-3 mb-3">
  <div class="row g-2">
    <div class="col-md-4">
      <label class="form-label">الفصل الدراسي</label>
      <select name="term" class="form-select" onchange="this.form.submit()">
        <option value="الفصل الأول" {% if term=='الفصل الأول' %}selected{% endif %}>الفصل الأول</option>
        <option value="الفصل الثاني" {% if term=='الفصل الثاني' %}selected{% endif %}>الفصل الثاني</option>
      </select>
    </div>
  </div>
</form>

<div class="row g-3 mb-3">
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold">{{ s.school_class.name if s.school_class else '-' }}</div><div class="text-muted small">الصف</div></div></div>
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold">{{ overall.total_score|round(1) }} / {{ overall.total_max|round(1) }}</div><div class="text-muted small">مجموع الدرجات</div></div></div>
  <div class="col-md-3"><div class="card p-3 text-center"><div class="fw-bold">{{ overall.percentage }}%</div><div class="text-muted small">النسبة الكلية</div></div></div>
  <div class="col-md-3">
    <div class="card p-3 text-center">
      <div class="fw-bold {{ 'text-success' if overall.status=='ناجح' else 'text-danger' }}">{{ overall.status }}</div>
      <div class="text-muted small">النتيجة النهائية</div>
    </div>
  </div>
</div>

<div class="card">
  <table class="table table-hover mb-0">
    <thead><tr><th>المادة</th><th>الدرجة</th><th>الدرجة القصوى</th><th>النسبة %</th><th>الحالة</th></tr></thead>
    <tbody>
      {% for r in rows %}
      <tr>
        <td>{{ r.subject }}</td>
        <td>{{ r.score|round(1) }}</td>
        <td>{{ r.max_score|round(1) }}</td>
        <td>{{ r.percentage }}%</td>
        <td><span class="badge bg-{{ 'success' if r.status=='ناجح' else 'danger' }}">{{ r.status }}</span></td>
      </tr>
      {% else %}
      <tr><td colspan="5" class="text-center text-muted py-4">لا توجد درجات مسجلة لهذا الفصل</td></tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endblock %}
""",
    "import_students.html": """{% extends 'base.html' %}
{% block title %}استيراد طلاب من Excel{% endblock %}
{% block content %}
<h3 class="mb-3"><i class="bi bi-file-earmark-arrow-up"></i> استيراد طلاب من ملف Excel</h3>

<div class="card p-4 mb-3">
  <form method="post" enctype="multipart/form-data">
    <div class="mb-3">
      <label class="form-label">اختر ملف Excel (.xlsx)</label>
      <input type="file" name="excel_file" accept=".xlsx" class="form-control" required>
    </div>
    <button class="btn btn-primary"><i class="bi bi-upload"></i> استيراد</button>
    <a href="{{ url_for('students_list') }}" class="btn btn-outline-secondary">إلغاء</a>
  </form>
</div>

<div class="card p-4">
  <h6 class="mb-3"><i class="bi bi-info-circle"></i> يجب أن يحتوي الملف على صف عناوين أولاً، بنفس هذا الترتيب:</h6>
  <div class="table-responsive">
    <table class="table table-sm table-bordered text-center mb-0">
      <thead><tr>
        {% for h in headers %}<th>{{ h }}</th>{% endfor %}
      </tr></thead>
      <tbody>
        <tr>
          <td>محمد أحمد</td><td>12345</td><td>ذكر</td><td>2010-05-01</td><td>الصف الأول أ</td>
          <td>الخرطوم</td><td>0900000000</td><td>أحمد محمد</td><td>0911111111</td><td>2025-09-01</td>
        </tr>
      </tbody>
    </table>
  </div>
  <p class="text-muted small mt-3 mb-0">
    ملاحظات: "الاسم الكامل" حقل إلزامي. عمود "الصف" يجب أن يطابق اسم صف موجود مسبقاً في النظام
    (وإلا سيُضاف الطالب بدون صف). الطلاب الذين لديهم "رقم وطني" مكرر يتم تجاوزهم تلقائياً.
  </p>
</div>
{% endblock %}
""",
}

app.jinja_loader = DictLoader(TEMPLATES)


# ---------------------------------------------------------------------------
# النماذج (Database Models)
# ---------------------------------------------------------------------------

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(150), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='admin')  # admin / teacher
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class SchoolClass(db.Model):
    """الصف / الفصل الدراسي"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)          # مثال: الصف الأول أ
    grade_level = db.Column(db.String(50), nullable=False)    # مثال: الصف الأول
    academic_year = db.Column(db.String(20), nullable=False, default='2025-2026')
    homeroom_teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=True)

    students = db.relationship('Student', backref='school_class', lazy=True)
    homeroom_teacher = db.relationship('Teacher', foreign_keys=[homeroom_teacher_id])

    @property
    def student_count(self):
        return len(self.students)


class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(150), nullable=False)
    phone = db.Column(db.String(30))
    email = db.Column(db.String(120))
    specialty = db.Column(db.String(100))       # التخصص
    hire_date = db.Column(db.Date, default=date.today)
    is_active = db.Column(db.Boolean, default=True)


class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    grade_level = db.Column(db.String(50))       # المرحلة المرتبطة بالمادة
    teacher_id = db.Column(db.Integer, db.ForeignKey('teacher.id'), nullable=True)

    teacher = db.relationship('Teacher')


class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(150), nullable=False)
    national_id = db.Column(db.String(50), unique=True)
    gender = db.Column(db.String(10))            # ذكر / أنثى
    birth_date = db.Column(db.Date)
    address = db.Column(db.String(255))
    phone = db.Column(db.String(30))
    parent_name = db.Column(db.String(150))
    parent_phone = db.Column(db.String(30))
    class_id = db.Column(db.Integer, db.ForeignKey('school_class.id'), nullable=True)
    enrollment_date = db.Column(db.Date, default=date.today)
    is_active = db.Column(db.Boolean, default=True)

    attendances = db.relationship('Attendance', backref='student', lazy=True, cascade='all, delete-orphan')
    grades = db.relationship('Grade', backref='student', lazy=True, cascade='all, delete-orphan')
    fee_payments = db.relationship('FeePayment', backref='student', lazy=True, cascade='all, delete-orphan')


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(20), nullable=False, default='حاضر')  # حاضر / غائب / متأخر / غياب بعذر
    notes = db.Column(db.String(255))

    __table_args__ = (db.UniqueConstraint('student_id', 'date', name='uq_student_date'),)


class Grade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    term = db.Column(db.String(50), nullable=False, default='الفصل الأول')
    exam_type = db.Column(db.String(50), nullable=False, default='اختبار نهائي')  # اختبار قصير / منتصف الفصل / نهائي
    score = db.Column(db.Float, nullable=False, default=0)
    max_score = db.Column(db.Float, nullable=False, default=100)

    subject = db.relationship('Subject')

    @property
    def percentage(self):
        if self.max_score:
            return round((self.score / self.max_score) * 100, 1)
        return 0


class FeePayment(db.Model):
    """الرسوم الدراسية (اختياري بسيط)"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.Date, default=date.today)
    description = db.Column(db.String(255))


# ---------------------------------------------------------------------------
# أدوات مساعدة للمصادقة والصلاحيات
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.path))
        if session.get('role') != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_globals():
    return dict(current_year=datetime.now().year, session=session, school_name=SCHOOL_NAME)


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# المسارات (Routes)
# ---------------------------------------------------------------------------

def register_routes(app, db):

    # -------------------------------------------------------------
    # المصادقة
    # -------------------------------------------------------------
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            user = User.query.filter_by(username=username, is_active=True).first()
            if user and user.check_password(password):
                session['user_id'] = user.id
                session['username'] = user.username
                session['full_name'] = user.full_name
                session['role'] = user.role
                flash('تم تسجيل الدخول بنجاح', 'success')
                next_url = request.args.get('next') or url_for('dashboard')
                return redirect(next_url)
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    def logout():
        session.clear()
        flash('تم تسجيل الخروج', 'info')
        return redirect(url_for('login'))

    # -------------------------------------------------------------
    # لوحة التحكم
    # -------------------------------------------------------------
    @app.route('/')
    @login_required
    def dashboard():
        stats = {
            'students': Student.query.filter_by(is_active=True).count(),
            'teachers': Teacher.query.filter_by(is_active=True).count(),
            'classes': SchoolClass.query.count(),
            'subjects': Subject.query.count(),
        }
        today = date.today()
        today_attendance = Attendance.query.filter_by(date=today).count()
        today_absent = Attendance.query.filter_by(date=today, status='غائب').count()
        recent_students = Student.query.order_by(Student.id.desc()).limit(5).all()
        classes = SchoolClass.query.all()
        return render_template('dashboard.html', stats=stats,
                                today_attendance=today_attendance,
                                today_absent=today_absent,
                                recent_students=recent_students,
                                classes=classes, today=today)

    # -------------------------------------------------------------
    # إدارة الطلاب
    # -------------------------------------------------------------
    @app.route('/students')
    @login_required
    def students_list():
        q = request.args.get('q', '').strip()
        class_id = request.args.get('class_id', '')
        query = Student.query
        if q:
            query = query.filter(Student.full_name.contains(q))
        if class_id:
            query = query.filter_by(class_id=class_id)
        students = query.order_by(Student.full_name).all()
        classes = SchoolClass.query.all()
        return render_template('students.html', students=students, classes=classes,
                                q=q, selected_class=class_id)

    @app.route('/students/add', methods=['GET', 'POST'])
    @login_required
    def student_add():
        classes = SchoolClass.query.all()
        if request.method == 'POST':
            s = Student(
                full_name=request.form['full_name'].strip(),
                national_id=request.form.get('national_id') or None,
                gender=request.form.get('gender'),
                birth_date=parse_date(request.form.get('birth_date')),
                address=request.form.get('address'),
                phone=request.form.get('phone'),
                parent_name=request.form.get('parent_name'),
                parent_phone=request.form.get('parent_phone'),
                class_id=request.form.get('class_id') or None,
                enrollment_date=parse_date(request.form.get('enrollment_date')) or date.today(),
            )
            db.session.add(s)
            db.session.commit()
            flash('تمت إضافة الطالب بنجاح', 'success')
            return redirect(url_for('students_list'))
        return render_template('student_form.html', student=None, classes=classes)

    @app.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
    @login_required
    def student_edit(student_id):
        s = Student.query.get_or_404(student_id)
        classes = SchoolClass.query.all()
        if request.method == 'POST':
            s.full_name = request.form['full_name'].strip()
            s.national_id = request.form.get('national_id') or None
            s.gender = request.form.get('gender')
            s.birth_date = parse_date(request.form.get('birth_date'))
            s.address = request.form.get('address')
            s.phone = request.form.get('phone')
            s.parent_name = request.form.get('parent_name')
            s.parent_phone = request.form.get('parent_phone')
            s.class_id = request.form.get('class_id') or None
            db.session.commit()
            flash('تم تحديث بيانات الطالب', 'success')
            return redirect(url_for('students_list'))
        return render_template('student_form.html', student=s, classes=classes)

    @app.route('/students/<int:student_id>/delete', methods=['POST'])
    @admin_required
    def student_delete(student_id):
        s = Student.query.get_or_404(student_id)
        db.session.delete(s)
        db.session.commit()
        flash('تم حذف الطالب', 'info')
        return redirect(url_for('students_list'))

    @app.route('/students/<int:student_id>')
    @login_required
    def student_profile(student_id):
        s = Student.query.get_or_404(student_id)
        grades = Grade.query.filter_by(student_id=student_id).all()
        attendances = Attendance.query.filter_by(student_id=student_id).order_by(Attendance.date.desc()).limit(30).all()
        total_present = Attendance.query.filter_by(student_id=student_id, status='حاضر').count()
        total_absent = Attendance.query.filter_by(student_id=student_id, status='غائب').count()
        payments = FeePayment.query.filter_by(student_id=student_id).all()
        total_paid = sum(p.amount for p in payments)
        return render_template('student_profile.html', s=s, grades=grades,
                                attendances=attendances, total_present=total_present,
                                total_absent=total_absent, payments=payments, total_paid=total_paid)

    register_teacher_routes(app, db)
    register_class_subject_routes(app, db)
    register_attendance_routes(app, db)
    register_grade_routes(app, db)
    register_fee_routes(app, db)
    register_user_routes(app, db)
    register_reports_routes(app, db)
    register_result_routes(app, db)
    register_excel_routes(app, db)


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


# ===================================================================
# المعلمون
# ===================================================================
def register_teacher_routes(app, db):

    @app.route('/teachers')
    @login_required
    def teachers_list():
        teachers = Teacher.query.order_by(Teacher.full_name).all()
        return render_template('teachers.html', teachers=teachers)

    @app.route('/teachers/add', methods=['GET', 'POST'])
    @admin_required
    def teacher_add():
        if request.method == 'POST':
            t = Teacher(
                full_name=request.form['full_name'].strip(),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                specialty=request.form.get('specialty'),
                hire_date=parse_date(request.form.get('hire_date')) or date.today(),
            )
            db.session.add(t)
            db.session.commit()
            flash('تمت إضافة المعلم بنجاح', 'success')
            return redirect(url_for('teachers_list'))
        return render_template('teacher_form.html', teacher=None)

    @app.route('/teachers/<int:teacher_id>/edit', methods=['GET', 'POST'])
    @admin_required
    def teacher_edit(teacher_id):
        t = Teacher.query.get_or_404(teacher_id)
        if request.method == 'POST':
            t.full_name = request.form['full_name'].strip()
            t.phone = request.form.get('phone')
            t.email = request.form.get('email')
            t.specialty = request.form.get('specialty')
            db.session.commit()
            flash('تم تحديث بيانات المعلم', 'success')
            return redirect(url_for('teachers_list'))
        return render_template('teacher_form.html', teacher=t)

    @app.route('/teachers/<int:teacher_id>/delete', methods=['POST'])
    @admin_required
    def teacher_delete(teacher_id):
        t = Teacher.query.get_or_404(teacher_id)
        db.session.delete(t)
        db.session.commit()
        flash('تم حذف المعلم', 'info')
        return redirect(url_for('teachers_list'))


# ===================================================================
# الصفوف والمواد
# ===================================================================
def register_class_subject_routes(app, db):

    @app.route('/classes')
    @login_required
    def classes_list():
        classes = SchoolClass.query.order_by(SchoolClass.name).all()
        return render_template('classes.html', classes=classes)

    @app.route('/classes/add', methods=['GET', 'POST'])
    @admin_required
    def class_add():
        teachers = Teacher.query.all()
        if request.method == 'POST':
            c = SchoolClass(
                name=request.form['name'].strip(),
                grade_level=request.form['grade_level'].strip(),
                academic_year=request.form.get('academic_year', '2025-2026'),
                homeroom_teacher_id=request.form.get('homeroom_teacher_id') or None,
            )
            db.session.add(c)
            db.session.commit()
            flash('تمت إضافة الصف بنجاح', 'success')
            return redirect(url_for('classes_list'))
        return render_template('class_form.html', school_class=None, teachers=teachers)

    @app.route('/classes/<int:class_id>/edit', methods=['GET', 'POST'])
    @admin_required
    def class_edit(class_id):
        c = SchoolClass.query.get_or_404(class_id)
        teachers = Teacher.query.all()
        if request.method == 'POST':
            c.name = request.form['name'].strip()
            c.grade_level = request.form['grade_level'].strip()
            c.academic_year = request.form.get('academic_year', c.academic_year)
            c.homeroom_teacher_id = request.form.get('homeroom_teacher_id') or None
            db.session.commit()
            flash('تم تحديث بيانات الصف', 'success')
            return redirect(url_for('classes_list'))
        return render_template('class_form.html', school_class=c, teachers=teachers)

    @app.route('/classes/<int:class_id>/delete', methods=['POST'])
    @admin_required
    def class_delete(class_id):
        c = SchoolClass.query.get_or_404(class_id)
        db.session.delete(c)
        db.session.commit()
        flash('تم حذف الصف', 'info')
        return redirect(url_for('classes_list'))

    @app.route('/subjects')
    @login_required
    def subjects_list():
        subjects = Subject.query.order_by(Subject.name).all()
        return render_template('subjects.html', subjects=subjects)

    @app.route('/subjects/add', methods=['GET', 'POST'])
    @admin_required
    def subject_add():
        teachers = Teacher.query.all()
        if request.method == 'POST':
            s = Subject(
                name=request.form['name'].strip(),
                grade_level=request.form.get('grade_level'),
                teacher_id=request.form.get('teacher_id') or None,
            )
            db.session.add(s)
            db.session.commit()
            flash('تمت إضافة المادة بنجاح', 'success')
            return redirect(url_for('subjects_list'))
        return render_template('subject_form.html', subject=None, teachers=teachers)

    @app.route('/subjects/<int:subject_id>/edit', methods=['GET', 'POST'])
    @admin_required
    def subject_edit(subject_id):
        s = Subject.query.get_or_404(subject_id)
        teachers = Teacher.query.all()
        if request.method == 'POST':
            s.name = request.form['name'].strip()
            s.grade_level = request.form.get('grade_level')
            s.teacher_id = request.form.get('teacher_id') or None
            db.session.commit()
            flash('تم تحديث بيانات المادة', 'success')
            return redirect(url_for('subjects_list'))
        return render_template('subject_form.html', subject=s, teachers=teachers)

    @app.route('/subjects/<int:subject_id>/delete', methods=['POST'])
    @admin_required
    def subject_delete(subject_id):
        s = Subject.query.get_or_404(subject_id)
        db.session.delete(s)
        db.session.commit()
        flash('تم حذف المادة', 'info')
        return redirect(url_for('subjects_list'))


# ===================================================================
# الحضور والغياب
# ===================================================================
def register_attendance_routes(app, db):

    @app.route('/attendance', methods=['GET', 'POST'])
    @login_required
    def attendance_page():
        classes = SchoolClass.query.all()
        class_id = request.args.get('class_id') or request.form.get('class_id')
        att_date = request.args.get('date') or request.form.get('date') or date.today().isoformat()

        if request.method == 'POST' and class_id:
            the_date = parse_date(att_date) or date.today()
            students = Student.query.filter_by(class_id=class_id, is_active=True).all()
            for s in students:
                status = request.form.get(f'status_{s.id}', 'حاضر')
                existing = Attendance.query.filter_by(student_id=s.id, date=the_date).first()
                if existing:
                    existing.status = status
                else:
                    db.session.add(Attendance(student_id=s.id, date=the_date, status=status))
            db.session.commit()
            flash('تم حفظ سجل الحضور بنجاح', 'success')
            return redirect(url_for('attendance_page', class_id=class_id, date=att_date))

        students = []
        existing_map = {}
        if class_id:
            students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.full_name).all()
            the_date = parse_date(att_date) or date.today()
            records = Attendance.query.filter_by(date=the_date).filter(
                Attendance.student_id.in_([s.id for s in students])).all() if students else []
            existing_map = {r.student_id: r.status for r in records}

        return render_template('attendance.html', classes=classes, students=students,
                                selected_class=class_id, selected_date=att_date,
                                existing_map=existing_map)


# ===================================================================
# الدرجات
# ===================================================================
def register_grade_routes(app, db):

    @app.route('/grades', methods=['GET', 'POST'])
    @login_required
    def grades_page():
        classes = SchoolClass.query.all()
        subjects = Subject.query.all()
        class_id = request.args.get('class_id') or request.form.get('class_id')
        subject_id = request.args.get('subject_id') or request.form.get('subject_id')
        term = request.args.get('term') or request.form.get('term', 'الفصل الأول')
        exam_type = request.args.get('exam_type') or request.form.get('exam_type', 'اختبار نهائي')

        if request.method == 'POST' and class_id and subject_id:
            students = Student.query.filter_by(class_id=class_id, is_active=True).all()
            for s in students:
                score_raw = request.form.get(f'score_{s.id}', '').strip()
                if score_raw == '':
                    continue
                try:
                    score = float(score_raw)
                except ValueError:
                    continue
                max_score = float(request.form.get('max_score', 100) or 100)
                existing = Grade.query.filter_by(student_id=s.id, subject_id=subject_id,
                                                  term=term, exam_type=exam_type).first()
                if existing:
                    existing.score = score
                    existing.max_score = max_score
                else:
                    db.session.add(Grade(student_id=s.id, subject_id=subject_id, term=term,
                                          exam_type=exam_type, score=score, max_score=max_score))
            db.session.commit()
            flash('تم حفظ الدرجات بنجاح', 'success')
            return redirect(url_for('grades_page', class_id=class_id, subject_id=subject_id,
                                     term=term, exam_type=exam_type))

        students = []
        existing_map = {}
        if class_id and subject_id:
            students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.full_name).all()
            records = Grade.query.filter_by(subject_id=subject_id, term=term, exam_type=exam_type).filter(
                Grade.student_id.in_([s.id for s in students])).all() if students else []
            existing_map = {r.student_id: r for r in records}

        return render_template('grades.html', classes=classes, subjects=subjects, students=students,
                                selected_class=class_id, selected_subject=subject_id,
                                term=term, exam_type=exam_type, existing_map=existing_map)


# ===================================================================
# الرسوم الدراسية
# ===================================================================
def register_fee_routes(app, db):

    @app.route('/fees')
    @login_required
    def fees_list():
        payments = FeePayment.query.order_by(FeePayment.payment_date.desc()).all()
        return render_template('fees.html', payments=payments)

    @app.route('/fees/add', methods=['GET', 'POST'])
    @admin_required
    def fee_add():
        students = Student.query.filter_by(is_active=True).order_by(Student.full_name).all()
        if request.method == 'POST':
            p = FeePayment(
                student_id=request.form['student_id'],
                amount=float(request.form['amount']),
                payment_date=parse_date(request.form.get('payment_date')) or date.today(),
                description=request.form.get('description'),
            )
            db.session.add(p)
            db.session.commit()
            flash('تم تسجيل الدفعة بنجاح', 'success')
            return redirect(url_for('fees_list'))
        return render_template('fee_form.html', students=students)


# ===================================================================
# المستخدمون (خاص بالمدير)
# ===================================================================
def register_user_routes(app, db):

    @app.route('/users')
    @admin_required
    def users_list():
        users = User.query.all()
        return render_template('users.html', users=users)

    @app.route('/users/add', methods=['GET', 'POST'])
    @admin_required
    def user_add():
        teachers = Teacher.query.all()
        if request.method == 'POST':
            u = User(
                username=request.form['username'].strip(),
                full_name=request.form['full_name'].strip(),
                role=request.form.get('role', 'teacher'),
                teacher_id=request.form.get('teacher_id') or None,
            )
            u.set_password(request.form['password'])
            db.session.add(u)
            db.session.commit()
            flash('تمت إضافة المستخدم بنجاح', 'success')
            return redirect(url_for('users_list'))
        return render_template('user_form.html', teachers=teachers)

    @app.route('/users/<int:user_id>/delete', methods=['POST'])
    @admin_required
    def user_delete(user_id):
        if user_id == session.get('user_id'):
            flash('لا يمكنك حذف حسابك الحالي', 'warning')
            return redirect(url_for('users_list'))
        u = User.query.get_or_404(user_id)
        db.session.delete(u)
        db.session.commit()
        flash('تم حذف المستخدم', 'info')
        return redirect(url_for('users_list'))


# ===================================================================
# التقارير
# ===================================================================
def register_reports_routes(app, db):

    @app.route('/reports/class/<int:class_id>')
    @login_required
    def report_class(class_id):
        c = SchoolClass.query.get_or_404(class_id)
        students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.full_name).all()
        report_rows = []
        for s in students:
            grades = Grade.query.filter_by(student_id=s.id).all()
            avg = round(sum(g.percentage for g in grades) / len(grades), 1) if grades else None
            present = Attendance.query.filter_by(student_id=s.id, status='حاضر').count()
            absent = Attendance.query.filter_by(student_id=s.id, status='غائب').count()
            report_rows.append({'student': s, 'avg': avg, 'present': present, 'absent': absent})
        return render_template('report_class.html', school_class=c, rows=report_rows)


# ===================================================================
# نتيجة الطالب (عرض + طباعة PDF)
# ===================================================================
PASS_PERCENT = 50  # نسبة النجاح الافتراضية لكل مادة


def compute_student_result(student_id, term):
    """يجمع كل درجات الطالب لهذا الفصل، مادة مادة، ويحسب النتيجة النهائية"""
    grades = Grade.query.filter_by(student_id=student_id, term=term).all()
    by_subject = {}
    for g in grades:
        by_subject.setdefault(g.subject_id, []).append(g)

    rows = []
    total_score = 0.0
    total_max = 0.0
    failed_subjects = 0
    for subject_id, glist in by_subject.items():
        subject = glist[0].subject
        s_score = sum(g.score for g in glist)
        s_max = sum(g.max_score for g in glist)
        pct = round((s_score / s_max) * 100, 1) if s_max else 0
        status = 'ناجح' if pct >= PASS_PERCENT else 'راسب'
        if status == 'راسب':
            failed_subjects += 1
        rows.append({
            'subject': subject.name,
            'score': s_score,
            'max_score': s_max,
            'percentage': pct,
            'status': status,
            'details': glist,
        })
        total_score += s_score
        total_max += s_max

    rows.sort(key=lambda r: r['subject'])
    overall_pct = round((total_score / total_max) * 100, 1) if total_max else 0
    overall_status = 'ناجح' if (overall_pct >= PASS_PERCENT and failed_subjects == 0) else 'راسب'

    overall = {
        'total_score': total_score,
        'total_max': total_max,
        'percentage': overall_pct,
        'status': overall_status,
        'failed_subjects': failed_subjects,
    }
    return rows, overall


def register_result_routes(app, db):

    @app.route('/students/<int:student_id>/result')
    @login_required
    def student_result(student_id):
        s = Student.query.get_or_404(student_id)
        term = request.args.get('term', 'الفصل الأول')
        rows, overall = compute_student_result(student_id, term)
        return render_template('student_result.html', s=s, rows=rows, overall=overall,
                                term=term, school_name=SCHOOL_NAME)

    @app.route('/students/<int:student_id>/result/pdf')
    @login_required
    def student_result_pdf(student_id):
        s = Student.query.get_or_404(student_id)
        term = request.args.get('term', 'الفصل الأول')
        rows, overall = compute_student_result(student_id, term)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
                                 topMargin=15 * mm, bottomMargin=15 * mm)

        title_style = ParagraphStyle('title', fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER,
                                      spaceAfter=6, leading=22)
        sub_style = ParagraphStyle('sub', fontName=PDF_FONT, fontSize=12, alignment=TA_CENTER,
                                    spaceAfter=4, leading=18)
        info_style = ParagraphStyle('info', fontName=PDF_FONT, fontSize=11, alignment=TA_RIGHT,
                                     spaceAfter=3, leading=16)

        elements = []
        elements.append(Paragraph(ar(SCHOOL_NAME), title_style))
        elements.append(Paragraph(ar('كشف نتيجة الطالب'), sub_style))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(ar(f'اسم الطالب: {s.full_name}'), info_style))
        elements.append(Paragraph(ar(f"الصف: {s.school_class.name if s.school_class else '-'}"), info_style))
        elements.append(Paragraph(ar(f'الفصل الدراسي: {term}'), info_style))
        elements.append(Spacer(1, 10))

        table_data = [[ar('الحالة'), ar('النسبة %'), ar('الدرجة القصوى'), ar('الدرجة'), ar('المادة')]]
        for r in rows:
            table_data.append([
                ar(r['status']), f"{r['percentage']}%", f"{r['max_score']:g}",
                f"{r['score']:g}", ar(r['subject']),
            ])
        if not rows:
            table_data.append([ar('-'), ar('-'), ar('-'), ar('-'), ar('لا توجد درجات مسجلة لهذا الفصل')])

        tbl = Table(table_data, colWidths=[25 * mm, 25 * mm, 30 * mm, 25 * mm, 55 * mm])
        tbl.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2a3a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f9')]),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 14))

        summary_style = ParagraphStyle('summary', fontName=PDF_FONT, fontSize=12, alignment=TA_RIGHT,
                                        spaceAfter=4, leading=18)
        elements.append(Paragraph(ar(f"مجموع الدرجات: {overall['total_score']:g} / {overall['total_max']:g}"), summary_style))
        elements.append(Paragraph(ar(f"النسبة المئوية الكلية: {overall['percentage']}%"), summary_style))
        elements.append(Paragraph(ar(f"عدد المواد الراسب فيها: {overall['failed_subjects']}"), summary_style))
        elements.append(Paragraph(ar(f"النتيجة النهائية: {overall['status']}"), summary_style))

        if not ARABIC_FONT_READY:
            warn_style = ParagraphStyle('warn', fontName='Helvetica', fontSize=8, alignment=TA_CENTER,
                                         textColor=colors.red, spaceBefore=10)
            elements.append(Paragraph(
                'Arabic font not found (Amiri-Regular.ttf) - place it next to flask_app.py for correct Arabic rendering.',
                warn_style))

        doc.build(elements)
        buf.seek(0)
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                          download_name=f"نتيجة_{s.full_name}.pdf")


# ===================================================================
# استيراد وتصدير بيانات الطلاب عبر ملفات إكسل (Excel)
# ===================================================================
EXCEL_HEADERS = [
    'الاسم الكامل', 'الرقم الوطني', 'الجنس', 'تاريخ الميلاد', 'الصف',
    'العنوان', 'هاتف الطالب', 'اسم ولي الأمر', 'هاتف ولي الأمر', 'تاريخ التسجيل',
]


def _cell_to_str(value):
    if value is None:
        return ''
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


def register_excel_routes(app, db):

    @app.route('/students/export')
    @login_required
    def students_export():
        students = Student.query.order_by(Student.full_name).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الطلاب'
        ws.sheet_view.rightToLeft = True
        ws.append(EXCEL_HEADERS)
        for s in students:
            ws.append([
                s.full_name, s.national_id or '', s.gender or '',
                s.birth_date.strftime('%Y-%m-%d') if s.birth_date else '',
                s.school_class.name if s.school_class else '',
                s.address or '', s.phone or '', s.parent_name or '', s.parent_phone or '',
                s.enrollment_date.strftime('%Y-%m-%d') if s.enrollment_date else '',
            ])
        for i, _ in enumerate(EXCEL_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          as_attachment=True, download_name='قائمة_الطلاب.xlsx')

    @app.route('/students/import', methods=['GET', 'POST'])
    @admin_required
    def students_import():
        if request.method == 'POST':
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                flash('الرجاء اختيار ملف إكسل', 'danger')
                return redirect(url_for('students_import'))
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash('تعذر قراءة الملف، تأكد أنه ملف Excel صحيح (xlsx)', 'danger')
                return redirect(url_for('students_import'))

            classes_by_name = {c.name: c for c in SchoolClass.query.all()}
            added, skipped = 0, 0
            rows_iter = ws.iter_rows(min_row=2, values_only=True)
            for row in rows_iter:
                if not row or not row[0]:
                    continue
                values = list(row) + [None] * (len(EXCEL_HEADERS) - len(row))
                full_name = _cell_to_str(values[0])
                if not full_name:
                    skipped += 1
                    continue
                national_id = _cell_to_str(values[1]) or None
                gender = _cell_to_str(values[2]) or None
                birth_date = parse_date(_cell_to_str(values[3])) if values[3] else None
                class_name = _cell_to_str(values[4])
                school_class = classes_by_name.get(class_name)
                address = _cell_to_str(values[5]) or None
                phone = _cell_to_str(values[6]) or None
                parent_name = _cell_to_str(values[7]) or None
                parent_phone = _cell_to_str(values[8]) or None
                enrollment_date = parse_date(_cell_to_str(values[9])) if values[9] else date.today()

                if national_id and Student.query.filter_by(national_id=national_id).first():
                    skipped += 1
                    continue

                s = Student(
                    full_name=full_name, national_id=national_id, gender=gender,
                    birth_date=birth_date, class_id=school_class.id if school_class else None,
                    address=address, phone=phone, parent_name=parent_name,
                    parent_phone=parent_phone, enrollment_date=enrollment_date,
                )
                db.session.add(s)
                added += 1

            db.session.commit()
            flash(f'تم استيراد {added} طالب بنجاح، وتم تجاوز {skipped} صف (بيانات ناقصة أو مكررة)', 'success')
            return redirect(url_for('students_list'))

        return render_template('import_students.html', headers=EXCEL_HEADERS)


register_routes(app, db)


def create_default_admin():
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', full_name='مدير النظام', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("تم إنشاء حساب المدير الافتراضي -> اسم المستخدم: admin | كلمة المرور: admin123")


with app.app_context():
    db.create_all()
    create_default_admin()


if __name__ == '__main__':
    # هذا الجزء يعمل فقط عند تشغيل الملف مباشرة كبرنامج سطح مكتب
    # (على سيرفر مثل PythonAnywhere، الملف يُستورد كـ WSGI ولا ينفّذ هذا الجزء إطلاقاً)
    import webview
    import threading
    import time

    def run_flask():
        # تشغيل السيرفر المحلي في مسار منفصل
        app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)

    # تشغيل Flask في خلفية البرنامج
    t = threading.Thread(target=run_flask)
    t.daemon = True
    t.start()

    # إعطاء السيرفر لحظة ليبدأ قبل فتح النافذة
    time.sleep(1)

    # فتح نافذة التطبيق المدمجة
    webview.create_window(SCHOOL_NAME, 'http://127.0.0.1:5000', width=1200, height=800)
    webview.start()